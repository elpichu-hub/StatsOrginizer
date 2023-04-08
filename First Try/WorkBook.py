from openpyxl import workbook, load_workbook
import openpyxl

# with an excel workbook created load it to 
# update it
wb = load_workbook('Work.xlsx')

# this selects the main sheet in the workbook
ws = wb.active

# deletes specific columns by index
# column A = index 1
ws.delete_cols(1)
ws.delete_cols(2)
ws.delete_cols(3)
ws.delete_cols(4)


# gives 'think' border to all cells
for row in ws.rows:
    for cell in row:
        # Aligns the cell's content to the center of the cell
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                             right=openpyxl.styles.Side(style='thin'),
                                             top=openpyxl.styles.Side(style='thin'),
                                             bottom=openpyxl.styles.Side(style='thin'))

# set all columns to 10 width
for cols in ws.columns:
    col_letter = cols[0].column_letter
    ws.column_dimensions[col_letter].width = 10

# set individual columns width
ws.column_dimensions['A'].width = 20
ws.column_dimensions['C'].width = 17

# Merge and Center cells E1 & F1
ws.merge_cells('E1:F1')
ws['E1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# This selects the first row's cells to provide font, background color
firstRow = ws[1]
for cell in firstRow:
    cell.font = openpyxl.styles.Font(size=14)
    cell.fill = openpyxl.styles.PatternFill(start_color='B7CEEC', end_color='B7CEEC', fill_type='solid')

# This will set the height of the first row only
ws.row_dimensions[1].height = 20


# change this to the word you want to search for
target_word = 'EndPoint' 

# This will delete everything below the target_word
# change max_col to the maximun column where the target_word
# should be found
for row in ws.iter_rows(min_row=1, max_col=7):
    if row[0].value == target_word:
        ws.delete_rows(row[0].row, ws.max_row)


# this will save the changes. 
# the work book needs to be closed 
# in order to save changes. 
wb.save('WorkDone.xlsx')