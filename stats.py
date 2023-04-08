from openpyxl import load_workbook
from openpyxl import utils, styles
import os
from datetime import datetime, time
import pandas as pd

def run_daily_stats():

    # this will make sure that no matter which extentions
    # the file is saved the program still runs
    # xls or xlsx
    try:
        df = pd.read_excel('stats.xls')
        df.to_excel('stats.xlsx', index=False)
    except Exception as e:
        print(f'An error occurred: {e}')
        
    # with an excel workbook created load it to 
    # update it. The file needs to have extension .xlsx,
    # .xls won't work
    work_book = load_workbook('stats.xlsx')

    # this selects the main sheet in the workbook
    work_sheet = work_book.active

    # Get the report Date from file without changes
    # and format the date to xx/xx/xxxx
    report_date_from_ICBM = work_sheet['F5'].value

    # Extract the start date from the string
    report_date_string = report_date_from_ICBM.split(" - ")[0]
    start_date = datetime.strptime(report_date_string, "%m/%d/%Y %I:%M:%S %p")

    # Format the start date as a string with only the date
    report_date = start_date.strftime("%m/%d/%Y")
    print(report_date + ' Report date')

    # This will loop through all the images
    # on the workbook and delete them all.
    for image in work_sheet._images:
        # Remove the image
        work_sheet._images.remove(image)


    # Remove rows 1, 11
    work_sheet.delete_rows(1, 11)
    work_sheet.delete_rows(2)
    work_sheet.delete_rows(4, 5)

    # set a starting points to start seaching
    # empty rows after such point
    start_row = 3

    # Loop through the rows in the worksheet
    for row in range(start_row, work_sheet.max_row + 1):
        # Check if all cells in the row are empty
        if all(cell.value is None for cell in work_sheet[row]):
            # An empty row was found
            empty_row = row
            print(f'Next Empty Row {empty_row}')
            work_sheet.delete_rows(empty_row, 6)
            break
    
    # change this to the word you want to search for
    target_phrase = 'Summaries Per User And Queue' 


    # This will delete everything below the target_word
    # change max_col to the maximun column where the target_word
    # should be found
    for row in work_sheet.iter_rows(min_row=1, max_col=7):
        if row[0].value == target_phrase:
            work_sheet.delete_rows(row[0].row, work_sheet.max_row)



    # At his point all the important data is all remaining 
    # from the rows. Columns Cleaning starts here.
    # ---------------------------------------------------------------------- #

    # Set the width of all columns to 10 
    for column in range(1, work_sheet.max_column + 1):
        column_letter = utils.get_column_letter(column)
        work_sheet.column_dimensions[column_letter].width = 12

    # Delete cols:
    work_sheet.delete_cols(2, 4)
    work_sheet.delete_cols(4, 6)
    work_sheet.delete_cols(5, 3)
    work_sheet.delete_cols(11, 3)


    # Delete all columns after handle time or 11
    num_cols = work_sheet.max_column - 11
    # Delete the columns
    if num_cols > 0:
        work_sheet.delete_cols(12, num_cols)

    # At this points only the necessary rows and columns
    # are displayed
    # ------------------------------------------------------------------------- #  

    # assign values to specific cells.
    work_sheet['A1'].value = f'Daily Stats {report_date}'
    work_sheet['B1'].value = 'Offered'
    work_sheet['B2'].value = '#'
    work_sheet['K1'].value = 'Handle Time'
    work_sheet['K2'].value = 'Average'


    # styling
    # Define the border style
    thin_border = styles.Border(
        left=styles.Side(style='thin'),
        right=styles.Side(style='thin'),
        top=styles.Side(style='thin'),
        bottom=styles.Side(style='thin')
    )




    # add width to specific columns
    work_sheet.column_dimensions['A'].width = 25
    work_sheet.column_dimensions['D'].width = 14
    work_sheet.column_dimensions['K'].width = 15

    # merge cells
    work_sheet.merge_cells('A1:A2')
    work_sheet.merge_cells('E1:F1')
    work_sheet.merge_cells('G1:H1')
    work_sheet.merge_cells('I1:J1')

    merged_range = 'A1:A2'

    # Create a Style object with centered alignment
    center_alignment = styles.Alignment(horizontal='center', vertical='center')

    # Loop over all rows and columns and apply the centered alignment
    for row in work_sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # Set the height of rows 1 and 2 to 20 and all other rows to 17
    for row in range(1, work_sheet.max_row + 1):
        if row == 1 or row == 2:
            work_sheet.row_dimensions[row].height = 20
            work_sheet.row_dimensions[row].font = styles.Font(size=14, bold=True)
        else:
            work_sheet.row_dimensions[row].height = 17
            work_sheet.row_dimensions[row].font = styles.Font(size=11)


    # Set the height of rows 1 and 2 to 20 and all other rows to 17
    for row in range(1, work_sheet.max_row + 1):
        if row == 1 or row == 2:
            work_sheet.row_dimensions[row].height = 25
            for cell in work_sheet[row]:
                cell.font = styles.Font(size=12, bold=True)
        else:
            work_sheet.row_dimensions[row].height = 17
            for cell in work_sheet[row]:
                cell.font = styles.Font(size=11)


    # this will loop thorugh all the rows from 4 to the last one
    # only rows with data will be looped through 
    new_data = []
    for row in range(4, work_sheet.max_row + 1):
        values = [cell.value for cell in work_sheet[row]]
        if all(value is None for value in values):
            continue  # skip empty rows
        new_data.append(values)

    # this will sort the data based on the names
    sorted_data = sorted(new_data, key=lambda x: x[0]) 

    # this gets the data sorted out based on the handle time average
    sorted_data_for_top_5 = sorted(new_data, key=lambda x: x[10]) 

    # this are the top 5
    top_5 = sorted_data_for_top_5[:5]

    # delete the data in the work_sheet and added 
    # sorted out
    for row in work_sheet.iter_rows(min_row=4):
        for cell in row:
            cell.value = None

    # I don't understand this lines quite well,
    # this will assigned all the sorted data
    # to each row and cell.
    for i, row in enumerate(sorted_data, start=4):
        for j, value in enumerate(row):
            work_sheet.cell(row=i, column=j+1, value=value)

    # this will make the 3rd row grey.
    grey = styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    for row in work_sheet.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.fill = grey

    green = styles.PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
    orange = styles.PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
    red = styles.PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')
    # on column K color based on conditions
    for cell in work_sheet['K']:
        try:
            time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
            # if 6:30 or less green
            if isinstance(time_obj, time) and time_obj < time(hour=0, minute=6, second=30):
                cell.fill = green
            if isinstance(time_obj, time) and time_obj < time(hour=0, minute=7, second=0) and time_obj > time(hour=0, minute=6, second=30):
                cell.fill = orange
            if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                cell.fill = red
        except ValueError:
            pass

    # more than 30 seconds will be red
    for cell in work_sheet['J']:
        try:
            time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
            # if 6:30 or less green
            if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                cell.fill = red
        except ValueError:
            pass

    # more than 30 seconds will be red
    for cell in work_sheet['H']:
        try:
            time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
            # if 6:30 or less green
            if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                cell.fill = red
        except ValueError:
            pass

    # more than 7 minutes will be red
    for cell in work_sheet['F']:
        try:
            time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
            # if 6:30 or less green
            if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                cell.fill = red
        except ValueError:
            pass



    blue = styles.PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
    # make row 1 and 2 blue
    for row in work_sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = blue

    # top 5 table creation starts here
    # ---------------------------------------- #
    work_sheet.merge_cells('N29:R30')
    work_sheet['N29'].value = 'Top 5!'
    work_sheet['N29'].alignment  = center_alignment
    work_sheet['N29'].font = styles.Font(size=14, bold=True)
    work_sheet['N29'].fill = blue
    work_sheet.merge_cells('N31:O32')
    work_sheet['N31'].value = f'Daily Stats {report_date}'
    work_sheet['N31'].alignment  = center_alignment 
    work_sheet['N31'].font = styles.Font(size=12, bold=True)

    work_sheet['P31'].value = 'Offered'
    work_sheet['P32'].value = '#'
    work_sheet['Q31'].value = 'Answered'
    work_sheet['Q32'].value = '#'
    work_sheet['R31'].value = 'Handle Time'
    work_sheet.column_dimensions['R'].width = 13
    work_sheet['R32'].value = 'Average'

    work_sheet.merge_cells('N33:O33')
    work_sheet.merge_cells('N34:O34')
    work_sheet.merge_cells('N35:O35')
    work_sheet.merge_cells('N36:O36')
    work_sheet.merge_cells('N37:O37')


    # Apply the border style to cells with data
    for row in work_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border


    # this is the list of all the merged cells ranges
    # I put them here so I can add border with the 
    # functions add_borders_to_merged_cells
    merged_ranges = ['A1:A2', 'E1:F1', 'G1:H1', 'I1:J1', 'N29:R30', 'N31:O32']

    # this will add borders to the ranges from merged_ranges
    def add_borders_to_merged_cells(merged_ranges):
        for merged_range in merged_ranges:
            for row in work_sheet[merged_range]:
                for cell in row:
                    cell.border = thin_border


    add_borders_to_merged_cells(merged_ranges)


    # loop through each row in the range N33 to R37
    for row in work_sheet['N33:R37']:
        # loop through each cell in the row and apply the border
        for cell in row:
            cell.border = thin_border


    # this will itarate through the top 5 and 
    # assign the values to the cells
    # above is a very repetive way to do the same
     # Write top 5 employees to the worksheet
    for i, row in enumerate(top_5[:5], start=33):
        work_sheet[f'N{i}'].value = row[0]
        work_sheet[f'P{i}'].value = row[1]
        work_sheet[f'Q{i}'].value = row[2]
        work_sheet[f'R{i}'].value = row[10]
        
        # center and align the content in the cell
        for col in ('N', 'P', 'Q', 'R'):
            cell = work_sheet[f'{col}{i}']
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')
    
    for row in range(31, 33):
        for col in ['N', 'O', 'P', 'Q', 'R']:
            cell = work_sheet[f'{col}{row}']
            cell.alignment = center_alignment
            cell.fill = blue
            cell.font = styles.Font(size=12, bold=True)

    
    # creates a date string to name the file
    date_for_file_name = start_date.strftime("%m.%d.%Y")

    # this will save the changes. 
    # The work book needs to be closed 
    # in order to save changes. 
    work_book.save(f'{date_for_file_name}.xlsx')


    # opens file after saved.
    os.startfile(f'{date_for_file_name}.xlsx')

    # deletes the statsNew workbook created at the start
    # with pandas to convert from xls to xlsx
    # os.remove('stats.xlsx')
    # os.remove('stats.xls')

